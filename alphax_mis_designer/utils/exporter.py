import frappe
from frappe import _
import io

@frappe.whitelist()
def export_canvas_excel(title: str, rows: list, meta: dict = None):
    if isinstance(rows, str):
        rows = frappe.parse_json(rows)
    if isinstance(meta, str):
        meta = frappe.parse_json(meta)

    try:
        import openpyxl
    except Exception:
        frappe.throw(_("openpyxl is required."))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Canvas"
    r = 1
    ws.cell(row=r, column=1, value=title or "AlphaX MIS Canvas"); r += 2

    if meta:
        for k in ["company","from_date","to_date","mode","dimensions","visual_template","layout_title"]:
            if meta.get(k):
                ws.cell(row=r, column=1, value=str(k))
                ws.cell(row=r, column=2, value=str(meta.get(k)))
                r += 1
        r += 1

    headers = ["Section","Row Key","Label","Actual","Budget","Variance"]
    for c,h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    r += 1

    for x in (rows or []):
        ws.cell(row=r, column=1, value=x.get("section"))
        ws.cell(row=r, column=2, value=x.get("row_key"))
        ws.cell(row=r, column=3, value=x.get("label"))
        ws.cell(row=r, column=4, value=float(x.get("actual") or 0))
        ws.cell(row=r, column=5, value=float(x.get("budget") or 0))
        ws.cell(row=r, column=6, value=float(x.get("variance") or 0))
        r += 1

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    fname = (title or "alphax_mis_canvas").replace(" ", "_") + ".xlsx"
    filedoc = frappe.get_doc({
        "doctype":"File",
        "file_name": fname,
        "is_private": 1,
        "content": bio.getvalue()
    }).insert(ignore_permissions=True)

    return {"file_url": filedoc.file_url, "file_name": filedoc.file_name}
