import re
import openpyxl

MONTH_RE = re.compile(r"\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[\s\-_/]*('?\d{2,4})\b", re.IGNORECASE)

def _to_col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def _col_letter_to_num(col: str) -> int:
    col = (col or "").strip().upper()
    n = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            continue
        n = n * 26 + (ord(ch) - 64)
    return n

def _is_month_label(v) -> bool:
    if v is None:
        return False
    if hasattr(v, "strftime"):
        return True
    if isinstance(v, str):
        return bool(MONTH_RE.search(v.strip()))
    return False

def detect_header_row(ws, scan_rows: int = 40, scan_cols: int = 120, min_months: int = 3):
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        month_cols = []
        for c in range(1, min(scan_cols, ws.max_column) + 1):
            if _is_month_label(ws.cell(r, c).value):
                month_cols.append(c)
        month_cols = sorted(set(month_cols))
        if len(month_cols) >= min_months:
            step = (month_cols[1] - month_cols[0]) if len(month_cols) >= 2 else 1
            step = step or 1
            return {"row": r, "start_col": min(month_cols), "month_cols": month_cols, "step": step}
    return None

def suggest_row_mappings(
    xlsx_path: str,
    sheet_name: str | None = None,
    max_rows: int = 300,
    header_row: int | None = None,
    first_month_col: str | None = None,
    period_col_step: int | None = None,
    label_col: int | None = None,
):
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets[0]

    if header_row and first_month_col:
        hr = int(header_row)
        start_col = _col_letter_to_num(first_month_col)
        step = int(period_col_step or 1)
        lbl_col = int(label_col or 1)
        if start_col <= 0:
            return {"sheet": ws.title, "error": "Invalid first_month_col", "suggestions": []}
    else:
        header = detect_header_row(ws)
        if not header:
            return {"sheet": ws.title, "error": "Could not detect month header row", "suggestions": []}
        hr = header["row"]
        start_col = header["start_col"]
        step = header["step"]
        lbl_col = 1

    start_letter = _to_col_letter(start_col)
    suggestions = []
    for r in range(hr + 1, min(ws.max_row, hr + max_rows) + 1):
        label = None
        # label_col override
        cols_to_try = [lbl_col] if (lbl_col and 1 <= lbl_col <= 15) else []
        cols_to_try += [c for c in range(1, 6) if c not in cols_to_try]
        for c in cols_to_try:
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() and len(v.strip()) <= 160:
                label = v.strip()
                break
        if not label:
            continue
        v1 = ws.cell(r, start_col).value
        v2 = ws.cell(r, start_col + step).value if (start_col + step) <= ws.max_column else None
        if v1 is None and v2 is None:
            continue
        row_key = re.sub(r"[^A-Za-z0-9]+", "_", label).strip("_")[:60] or f"ROW_{r}"
        base_cell = f"{ws.title}!{start_letter}{r}"
        suggestions.append({"row_key": row_key, "label": label, "base_cell": base_cell, "period_col_step": step})

    return {"sheet": ws.title, "header_row": hr, "first_month_col": start_letter, "period_col_step": step, "suggestions": suggestions}
