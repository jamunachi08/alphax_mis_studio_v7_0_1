import io, json
import frappe
from frappe import _
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from alphax_mis_designer.utils.runner import run_template

def _apply_style(cell, s: dict):
    if not s:
        return
    try:
        nf = s.get("number_format")
        if nf:
            cell.number_format = nf
        f = s.get("font") or {}
        cell.font = Font(
            name=f.get("name"), size=f.get("sz"),
            bold=f.get("b"), italic=f.get("i"),
            underline="single" if f.get("u") else None,
            color=f.get("color"),
        )
        a = s.get("alignment") or {}
        cell.alignment = Alignment(horizontal=a.get("horizontal"), vertical=a.get("vertical"), wrap_text=a.get("wrap_text"))
        fl = s.get("fill") or {}
        fg = fl.get("fgColor")
        if fg:
            cell.fill = PatternFill(patternType=fl.get("patternType") or "solid", fgColor=fg)
        b = s.get("border") or {}
        def sd(style): return Side(style=style) if style else Side(style=None)
        cell.border = Border(left=sd(b.get("left")), right=sd(b.get("right")), top=sd(b.get("top")), bottom=sd(b.get("bottom")))
    except Exception:
        pass

def export_filled_template(template_name: str, filters: dict):
    tpl = frappe.get_doc("MIS Report Template", template_name)
    if not tpl.layout_json:
        frappe.throw(_("Attach XLSX and click Import From Excel first."))
    layout = json.loads(tpl.layout_json)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sh in layout.get("sheets", []):
        ws = wb.create_sheet(title=sh.get("title") or "Sheet1")
        for r, h in (sh.get("row_heights") or {}).items():
            ws.row_dimensions[int(r)].height = float(h)
        for col, w in (sh.get("col_widths") or {}).items():
            ws.column_dimensions[col].width = float(w)
        for m in (sh.get("merges") or []):
            ws.merge_cells(m)
        for addr, meta in (sh.get("cells") or {}).items():
            c = ws[addr]
            c.value = meta.get("v")
            _apply_style(c, meta.get("s") or {})
    computed = run_template(template_name=template_name, filters=filters, preview_only=True)
    for write in computed.get("writes", []):
        cell_ref = write["cell"]
        val = write["value"]
        if "!" in cell_ref:
            sname, addr = cell_ref.split("!", 1)
            if sname in wb.sheetnames:
                wb[sname][addr].value = val
        else:
            wb[wb.sheetnames[0]][cell_ref].value = val
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), f"{template_name}-MIS.xlsx"
