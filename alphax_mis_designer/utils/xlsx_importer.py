import openpyxl

def _cell_style_dict(c) -> dict:
    try:
        font = c.font
        fill = c.fill
        align = c.alignment
        border = c.border
        return {
            "number_format": c.number_format,
            "font": {
                "name": font.name, "sz": font.sz, "b": bool(font.b),
                "i": bool(font.i), "u": bool(font.u),
                "color": getattr(font.color, "rgb", None),
            },
            "fill": {
                "patternType": getattr(fill, "patternType", None),
                "fgColor": getattr(getattr(fill, "fgColor", None), "rgb", None),
            },
            "alignment": {
                "horizontal": align.horizontal, "vertical": align.vertical,
                "wrap_text": bool(align.wrap_text),
            },
            "border": {
                "left": getattr(border.left, "style", None),
                "right": getattr(border.right, "style", None),
                "top": getattr(border.top, "style", None),
                "bottom": getattr(border.bottom, "style", None),
            },
        }
    except Exception:
        return {}

def import_xlsx(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=False)
    out = {"sheets": []}
    for ws in wb.worksheets:
        sheet = {
            "title": ws.title,
            "row_heights": {},
            "col_widths": {},
            "merges": [],
            "cells": {},
            "max_row": ws.max_row,
            "max_col": ws.max_column,
        }
        for r in range(1, ws.max_row + 1):
            dim = ws.row_dimensions.get(r)
            if dim and dim.height:
                sheet["row_heights"][str(r)] = float(dim.height)
        for col_letter, dim in ws.column_dimensions.items():
            if dim and dim.width:
                sheet["col_widths"][col_letter] = float(dim.width)
        for m in ws.merged_cells.ranges:
            sheet["merges"].append(str(m))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for c in row:
                if c.value is None and not c.has_style:
                    continue
                sheet["cells"][c.coordinate] = {"v": c.value, "s": _cell_style_dict(c)}
        out["sheets"].append(sheet)
    return out
