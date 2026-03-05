import frappe
from frappe import _
from difflib import SequenceMatcher
import io

def _get_file_content(file_url: str) -> bytes:
    # file_url can be /files/.. or file doc name
    if not file_url:
        return b""
    if file_url.startswith("/files/") or file_url.startswith("/private/files/"):
        # from File doctype
        f = frappe.get_all("File", filters={"file_url": file_url}, fields=["name"], limit=1)
        if f:
            file_doc = frappe.get_doc("File", f[0].name)
            return file_doc.get_content()
    # else treat as File name
    if frappe.db.exists("File", file_url):
        return frappe.get_doc("File", file_url).get_content()
    return b""

def _ratio(a: str, b: str) -> float:
    return SequenceMatcher(None, (a or "").lower(), (b or "").lower()).ratio()

def suggest_accounts(company: str, label: str, limit: int = 5):
    label = (label or "").strip()
    if not label:
        return []

    # fetch candidate accounts (leaf only) limited
    rows = frappe.get_all("Account",
        filters={"company": company, "is_group": 0, "disabled": 0},
        fields=["name","account_name"],
        limit_page_length=5000
    )

    scored = []
    for r in rows:
        name = r.get("name")
        acc_name = r.get("account_name") or ""
        score = max(_ratio(label, name), _ratio(label, acc_name))
        # bonus if label tokens appear in acc name
        toks = [t for t in label.lower().split() if len(t) >= 3]
        bonus = sum(1 for t in toks if t in (name or "").lower() or t in acc_name.lower()) * 0.03
        score = score + bonus
        if score >= 0.45:
            scored.append((score, name))
    scored.sort(reverse=True, key=lambda x: x[0])
    return [x[1] for x in scored[: int(limit or 5)]]

def analyze_excel(file_url: str, sheet_name: str = None, label_column: int = 1, start_row: int = 1, max_rows: int = 500):
    content = _get_file_content(file_url)
    if not content:
        frappe.throw(_("Unable to read Excel file."))

    try:
        import openpyxl
    except Exception:
        frappe.throw(_("openpyxl not available on server."))

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    sheet_name = sheet.title

    labels = []
    col = int(label_column or 1)
    sr = int(start_row or 1)

    for r in range(sr, min(sr + int(max_rows or 500), sheet.max_row + 1)):
        v = sheet.cell(row=r, column=col).value
        if v is None:
            continue
        txt = str(v).strip()
        if not txt:
            continue
        # skip obvious headers
        if txt.lower() in ("account", "accounts", "particulars", "description"):
            continue
        labels.append({"row": r, "label": txt})

    return {"sheet_name": sheet_name, "label_column": col, "start_row": sr, "labels": labels, "total_labels": len(labels)}

@frappe.whitelist()
def analyze_excel_whitelist(excel_file: str, sheet_name: str = None, label_column: int = 1, start_row: int = 1):
    return analyze_excel(excel_file, sheet_name, label_column, start_row)

@frappe.whitelist()
def create_visual_template_from_excel(company: str, title: str, analysis: dict, section_name: str = "Imported"):
    if isinstance(analysis, str):
        analysis = frappe.parse_json(analysis)

    labels = (analysis or {}).get("labels") or []
    if not labels:
        frappe.throw(_("No labels found in analysis."))

    vt = frappe.new_doc("MIS Visual Template")
    vt.template_title = title
    vt.description = f"Auto-created from Excel: {analysis.get('sheet_name')}"
    vt.default_mode = "Both"

    # create rows with mapping suggestions
    # first heading row for section
    vt.append("rows", {
        "section": section_name,
        "row_key": frappe.scrub(section_name),
        "row_type": "Heading",
        "label": section_name,
        "indent": 0,
        "is_bold": 1,
        "sign_factor": 0
    })

    for item in labels:
        lbl = item.get("label")
        # Suggest accounts (store as newline list, user can refine)
        sugg = suggest_accounts(company, lbl, limit=5)
        vt.append("rows", {
            "section": section_name,
            "row_key": frappe.scrub(lbl)[:60],
            "row_type": "Account",
            "label": lbl,
            "accounts": "\n".join(sugg) if sugg else "",
            "indent": 0,
            "is_bold": 0,
            "sign_factor": 1
        })

    vt.insert(ignore_permissions=True)
    return {"visual_template": vt.name}
