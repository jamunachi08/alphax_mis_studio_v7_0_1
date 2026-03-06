import json
import frappe
from frappe.model.document import Document
from frappe import _
from frappe.utils.file_manager import get_file
from openpyxl import load_workbook

def _json_safe(obj):
    rgb = getattr(obj, "rgb", None)
    if rgb:
        return str(rgb)
    for attr in ("indexed", "theme", "tint", "type"):
        val = getattr(obj, attr, None)
        if val is not None:
            try:
                return int(val)
            except Exception:
                return str(val)
    try:
        return str(obj)
    except Exception:
        return None

def _style_to_dict(cell):
    s = {}
    try:
        if cell.font:
            s["font"] = {
                "name": cell.font.name,
                "sz": cell.font.sz,
                "b": bool(cell.font.b),
                "i": bool(cell.font.i),
                "u": str(cell.font.u) if cell.font.u else None,
                "color": getattr(cell.font.color, "rgb", None) if cell.font.color else None,
            }
        if cell.fill and getattr(cell.fill, "fgColor", None):
            s["fill"] = {"fgColor": getattr(cell.fill.fgColor, "rgb", None)}
        if cell.alignment:
            s["align"] = {
                "h": cell.alignment.horizontal,
                "v": cell.alignment.vertical,
                "wrap": bool(cell.alignment.wrap_text),
            }
        if cell.number_format:
            s["numfmt"] = cell.number_format
        if cell.border:
            s["border"] = {
                "l": bool(cell.border.left and cell.border.left.style),
                "r": bool(cell.border.right and cell.border.right.style),
                "t": bool(cell.border.top and cell.border.top.style),
                "b": bool(cell.border.bottom and cell.border.bottom.style),
            }
    except Exception:
        pass
    return s

class MISVisualTemplate(Document):

    @frappe.whitelist()
    def import_layout(self):
        if not self.template_xlsx:
            frappe.throw(_("Please attach an XLSX file"))

        file_path = get_file(self.template_xlsx)[1]
        wb = load_workbook(filename=file_path, data_only=False)
        ws = wb[self.sheet_name] if self.sheet_name and self.sheet_name in wb.sheetnames else wb.worksheets[0]

        merges = [str(rng) for rng in ws.merged_cells.ranges]

        col_widths = {}
        for col_letter, dim in ws.column_dimensions.items():
            if dim and dim.width:
                col_widths[col_letter] = float(dim.width)

        row_heights = {}
        for row_idx, dim in ws.row_dimensions.items():
            if dim and dim.height:
                row_heights[str(row_idx)] = float(dim.height)

        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        cells = []
        for r in range(1, max_row+1):
            row = []
            for c in range(1, max_col+1):
                cell = ws.cell(row=r, column=c)
                row.append({
                    "r": r, "c": c,
                    "v": cell.value,
                    "t": "value",
                    "style": _style_to_dict(cell),
                })
            cells.append(row)

        layout = {
            "sheet": ws.title,
            "max_row": max_row,
            "max_col": max_col,
            "merges": merges,
            "col_widths": col_widths,
            "row_heights": row_heights,
            "cells": cells,
        }
        self.layout_json = json.dumps(layout, ensure_ascii=False, default=_json_safe)
        self.save(ignore_permissions=True)
        return {"ok": True, "rows": max_row, "cols": max_col, "sheet": ws.title}
