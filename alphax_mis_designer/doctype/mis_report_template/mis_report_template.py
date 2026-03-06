import json
import re
import frappe
from frappe.model.document import Document
from frappe import _
from frappe.utils.file_manager import get_file
from openpyxl import load_workbook

def _json_safe(obj):
    rgb = getattr(obj, "rgb", None)
    if rgb:
        return str(rgb)
    for attr in ("indexed", "theme", "tint", "type"):
        val = getattr(obj, attr, None)
        if val is not None:
            try:
                return int(val)
            except Exception:
                return str(val)
    try:
        return str(obj)
    except Exception:
        return None

class MISReportTemplate(Document):

    @frappe.whitelist()
    def import_from_excel(self):
        file_url = self.template_xlsx
        if getattr(self, "visual_template", None):
            vt = frappe.get_doc("MIS Visual Template", self.visual_template)
            if vt.layout_json:
                self.layout_json = vt.layout_json
                self.save(ignore_permissions=True)
            file_url = vt.template_xlsx or file_url

        if not file_url:
            frappe.throw(_("Attach an XLSX or select a MIS Visual Template"))

        file_path = get_file(file_url)[1]
        wb = load_workbook(filename=file_path, data_only=False)
        ws = wb.worksheets[0]

        merges = [str(rng) for rng in ws.merged_cells.ranges]
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        cells = []
        for r in range(1, max_row + 1):
            row = []
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                row.append({"r": r, "c": c, "v": cell.value, "t": "value"})
            cells.append(row)

        layout = {"sheet": ws.title, "max_row": max_row, "max_col": max_col, "merges": merges, "cells": cells}
        self.layout_json = json.dumps(layout, ensure_ascii=False, default=_json_safe)
        self.save(ignore_permissions=True)
        return {"ok": True, "rows": max_row, "cols": max_col}

    @frappe.whitelist()
    def import_map_sheet(self):
        file_url = self.template_xlsx
        if getattr(self, "visual_template", None):
            vt = frappe.get_doc("MIS Visual Template", self.visual_template)
            file_url = vt.template_xlsx or file_url

        if not file_url:
            frappe.throw(_("Attach an XLSX or select a MIS Visual Template"))

        map_sheet = (self.map_sheet_name or "MAP").strip() or "MAP"
        file_path = get_file(file_url)[1]
        wb = load_workbook(filename=file_path, data_only=True)
        if map_sheet not in wb.sheetnames:
            frappe.throw(_("MAP sheet not found: {0}").format(map_sheet))

        ws = wb[map_sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            frappe.throw(_("MAP sheet is empty"))

        header_idx = None
        header = []
        for idx, row in enumerate(rows[:15], start=1):
            norm = [str(x).strip().lower() if x is not None else "" for x in row]
            joined = " | ".join(norm)
            if "classification" in joined or "p&l" in joined or "account" in joined or "label" in joined:
                header_idx = idx
                header = norm
                break
        if not header_idx:
            for idx, row in enumerate(rows, start=1):
                norm = [str(x).strip().lower() if x is not None else "" for x in row]
                if any(norm):
                    header_idx = idx
                    header = norm
                    break

        def _find_col(cands):
            for i, h in enumerate(header):
                for c in cands:
                    if c in h:
                        return i
            return None

        acc_idx = _find_col(["account", "gl", "ledger"])
        label_idx = _find_col(["classification", "p&l", "label", "group", "report row"])
        if acc_idx is None or label_idx is None:
            non_empty = [i for i, h in enumerate(header) if h]
            if len(non_empty) >= 2:
                acc_idx = non_empty[0]
                label_idx = non_empty[-1]
            else:
                frappe.throw(_("Could not identify Account and Label columns in MAP sheet"))

        mapping = {}
        raw = []
        for row in rows[header_idx:]:
            vals = list(row)
            acc = vals[acc_idx] if acc_idx < len(vals) else None
            lbl = vals[label_idx] if label_idx < len(vals) else None
            acc = str(acc).strip() if acc is not None else ""
            lbl = str(lbl).strip() if lbl is not None else ""
            if not acc or not lbl:
                continue
            if acc.lower() in ("account", "account name", "gl account") or "classification" in lbl.lower():
                continue
            mapping.setdefault(lbl, []).append(acc)
            raw.append({"account": acc, "label": lbl})

        existing = list(self.row_mappings or [])
        keep = []
        for r in existing:
            if getattr(r, "formula", "") != "__AUTO_MAP__":
                keep.append({
                    "row_no": getattr(r, "row_no", None),
                    "row_type": getattr(r, "row_type", None),
                    "label": getattr(r, "label", ""),
                    "accounts": getattr(r, "accounts", ""),
                    "account_group": getattr(r, "account_group", ""),
                    "sign": getattr(r, "sign", "+1"),
                    "formula": getattr(r, "formula", ""),
                    "row_key": getattr(r, "row_key", ""),
                    "total_mode": getattr(r, "total_mode", ""),
                    "includes_row_nos": getattr(r, "includes_row_nos", ""),
                })

        max_row_no = 0
        for r in keep:
            try:
                max_row_no = max(max_row_no, int(r.get("row_no") or 0))
            except Exception:
                pass

        auto_rows = []
        seq = max_row_no + 1
        for label, accounts in mapping.items():
            row_key = re.sub(r"[^A-Z0-9]+", "_", label.upper()).strip("_")
            auto_rows.append({
                "doctype": "MIS Row Mapping",
                "row_no": seq,
                "row_type": "Account",
                "label": label,
                "accounts": ", ".join(accounts),
                "sign": "+1",
                "row_key": row_key,
                "formula": "__AUTO_MAP__"
            })
            seq += 1

        self.set("row_mappings", [])
        for r in keep:
            self.append("row_mappings", r)
        for r in auto_rows:
            self.append("row_mappings", r)

        self.save(ignore_permissions=True)
        return {"ok": True, "map_sheet": map_sheet, "rows_created": len(auto_rows), "sample": raw[:10]}

    @frappe.whitelist()
    def render_html(self, filters=None):
        filters = json.loads(filters) if isinstance(filters, str) and filters else (filters or {})
        layout = json.loads(self.layout_json) if self.layout_json else {}
        rows = list(self.row_mappings or [])
        cols = list(self.column_mappings or [])

        row_map = {}
        row_key_map = {}
        for r in rows:
            if r.row_no:
                row_map[int(r.row_no)] = r
            if getattr(r, "row_key", None):
                row_key_map[str(r.row_key).strip().upper()] = int(r.row_no)

        col_keys = [(c.col_key or c.col_label or f"COL{i+1}").strip() for i, c in enumerate(cols)]

        values = {}
        for rno, rm in row_map.items():
            rt = (rm.row_type or "").strip()
            if rt not in ("Account", "Account Group"):
                continue
            accounts = _resolve_accounts(rm)
            sign = -1 if (rm.sign or "+1").strip() == "-1" else 1
            for i, cm in enumerate(cols):
                key = col_keys[i]
                mode = (cm.balance_mode or "Period Net").strip()
                values[(rno, key)] = float(_compute_balance(accounts, mode, filters) or 0.0) * sign

        max_row_no = max(row_map.keys()) if row_map else 0
        for rno in sorted(row_map.keys()):
            rm = row_map[rno]
            if (rm.row_type or "").strip() != "Total":
                continue
            total_mode = (getattr(rm, "total_mode", "") or "Sum Children").strip() or "Sum Children"
            if total_mode == "Sum Selected Rows":
                picks = _parse_row_nos(getattr(rm, "includes_row_nos", ""))
                for key in col_keys:
                    values[(rno, key)] = sum(values.get((pn, key), 0.0) for pn in picks)
            elif total_mode == "Tree Rollup":
                accounts = _accounts_under_group(rm.account_group) if getattr(rm, "account_group", None) else []
                for i, cm in enumerate(cols):
                    key = col_keys[i]
                    mode = (cm.balance_mode or "Period Net").strip()
                    values[(rno, key)] = float(_compute_balance(accounts, mode, filters) or 0.0)
            else:
                child_rows = []
                for rr in range(rno + 1, max_row_no + 1):
                    if rr not in row_map:
                        continue
                    rt2 = (row_map[rr].row_type or "").strip()
                    if rt2 == "Heading" and (row_map[rr].label or "").strip():
                        break
                    if rt2 == "Total":
                        break
                    if rt2 in ("Account", "Account Group", "Formula"):
                        child_rows.append(rr)
                for key in col_keys:
                    values[(rno, key)] = sum(values.get((cr, key), 0.0) for cr in child_rows)

        for rno in sorted(row_map.keys()):
            rm = row_map[rno]
            if (rm.row_type or "").strip() != "Formula":
                continue
            expr = (rm.formula or "").strip()
            if not expr or expr == "__AUTO_MAP__":
                continue
            for key in col_keys:
                values[(rno, key)] = float(_eval_formula(expr, key, values, row_key_map) or 0.0)

        for i, cm in enumerate(cols):
            vt = (cm.value_type or "Amount").strip()
            if vt == "Amount":
                continue
            key = col_keys[i]
            base_key = (getattr(cm, "compare_to_col_key", "") or "").strip()
            if not base_key:
                continue
            for rno in row_map.keys():
                a = float(values.get((rno, key), 0.0))
                b = float(values.get((rno, base_key), 0.0))
                if vt == "Variance":
                    values[(rno, key)] = a - b
                elif vt in ("Percent", "Growth"):
                    values[(rno, key)] = (a / b * 100.0) if b else 0.0

        html = _render_excel_like_table(layout, col_keys, row_map, values)
        return {"html": html}

def _parse_row_nos(txt):
    out = []
    for p in (txt or "").split(","):
        p = p.strip()
        if not p:
            continue
        try:
            out.append(int(p))
        except Exception:
            pass
    return out

def _resolve_accounts(rm):
    rt = (rm.row_type or "").strip()
    if rt == "Account":
        return [a.strip() for a in (rm.accounts or "").split(",") if a.strip()]
    if rt == "Account Group" and getattr(rm, "account_group", None):
        return _accounts_under_group(rm.account_group)
    return []

def _accounts_under_group(group):
    acc = [group]
    q = [group]
    while q:
        parent = q.pop(0)
        children = frappe.get_all("Account", filters={"parent_account": parent}, pluck="name") or []
        for ch in children:
            if ch not in acc:
                acc.append(ch)
                q.append(ch)
    return acc

def _get_dept_fieldname():
    try:
        s = frappe.get_single("MIS Settings")
        if s and s.department_fieldname:
            return (s.department_fieldname or "cost_center").strip()
    except Exception:
        pass
    return "cost_center"

def _compute_balance(accounts, mode, filters):
    if not accounts:
        return 0.0
    company = filters.get("company")
    from_date = filters.get("from_date")
    to_date = filters.get("to_date")
    cost_center = filters.get("cost_center")
    project = filters.get("project")
    department = filters.get("department")

    params = {"accounts": tuple(accounts)}
    cond = ["account in %(accounts)s", "is_cancelled=0"]
    if company:
        cond.append("company=%(company)s"); params["company"] = company
    if cost_center:
        cond.append("cost_center=%(cost_center)s"); params["cost_center"] = cost_center
    if project:
        cond.append("project=%(project)s"); params["project"] = project
    if department:
        dept_field = _get_dept_fieldname()
        field = dept_field if dept_field in ("cost_center", "project") else "cost_center"
        cond.append(f"{field}=%(department)s"); params["department"] = department

    where = " and ".join(cond)
    expr = "sum(debit) - sum(credit)"

    if mode == "Opening":
        if not from_date:
            return 0.0
        params["from_date"] = from_date
        q = f"select {expr} from `tabGL Entry` where {where} and posting_date < %(from_date)s"
    elif mode == "Closing":
        if not to_date:
            return 0.0
        params["to_date"] = to_date
        q = f"select {expr} from `tabGL Entry` where {where} and posting_date <= %(to_date)s"
    elif mode == "YTD":
        if not to_date:
            return 0.0
        params["to_date"] = to_date
        q = f"select {expr} from `tabGL Entry` where {where} and posting_date <= %(to_date)s"
    elif mode == "MTD":
        if not to_date:
            return 0.0
        params["to_date"] = to_date
        q = f"""select {expr} from `tabGL Entry` where {where}
                and posting_date >= DATE_FORMAT(%(to_date)s,'%%Y-%%m-01') and posting_date <= %(to_date)s"""
    else:
        if not (from_date and to_date):
            return 0.0
        params["from_date"] = from_date
        params["to_date"] = to_date
        q = f"select {expr} from `tabGL Entry` where {where} and posting_date between %(from_date)s and %(to_date)s"
    amt = frappe.db.sql(q, params)[0][0]
    return float(amt or 0.0)

_ALLOWED_FUNCS = {"ABS": abs, "ROUND": round}

def _pct(a, b):
    try:
        return (float(a) / float(b) * 100.0) if float(b) else 0.0
    except Exception:
        return 0.0
_ALLOWED_FUNCS["PCT"] = _pct

def _row_value(ref, col_key, values, row_key_map):
    if isinstance(ref, int):
        return float(values.get((ref, col_key), 0.0))
    if isinstance(ref, str):
        rno = row_key_map.get(ref.strip().upper())
        if rno:
            return float(values.get((rno, col_key), 0.0))
    return 0.0

def _eval_formula(expr, col_key, values, row_key_map):
    s = re.sub(r"R\(\s*(\d+)\s*\)", lambda m: f"_row({int(m.group(1))})", expr)
    s = re.sub(r"K\(\s*['\"]([^'\"]+)['\"]\s*\)", lambda m: f"_row('{m.group(1)}')", s)
    if not re.fullmatch(r"[0-9A-Za-z_\s\+\-\*/\(\)\.,'\"]+", s):
        return 0.0
    def _row(x):
        return _row_value(x, col_key, values, row_key_map)
    env = {"__builtins__": {}, "_row": _row, **_ALLOWED_FUNCS}
    try:
        return float(eval(s, env, {}))
    except Exception:
        return 0.0

def _render_excel_like_table(layout, col_keys, row_map, values):
    max_row = int(layout.get("max_row") or 1)
    max_col = int(layout.get("max_col") or 1)
    cells = layout.get("cells") or []

    skip = set()
    spans = {}
    for m in (layout.get("merges") or []):
        try:
            a, b = m.split(":")
            import re as _re
            def rc(addr):
                col = _re.sub(r"[^A-Z]", "", addr)
                row = int(_re.sub(r"[^0-9]", "", addr))
                n = 0
                for ch in col:
                    n = n * 26 + (ord(ch) - 64)
                return row, n
            r1, c1 = rc(a); r2, c2 = rc(b)
            spans[(r1, c1)] = (r2 - r1 + 1, c2 - c1 + 1)
            for rr in range(r1, r2 + 1):
                for cc in range(c1, c2 + 1):
                    if (rr, cc) != (r1, c1):
                        skip.add((rr, cc))
        except Exception:
            continue

    start_value_col = max_col - len(col_keys) + 1 if col_keys else max_col + 1
    def fmt(v):
        try:
            return f"{float(v):,.2f}"
        except Exception:
            return "" if v is None else str(v)

    out = ["<div class='alphax-mis-excel'>",
           "<style>.alphax-mis-excel table{border-collapse:collapse;width:100%;} .alphax-mis-excel td{border:1px solid #ddd;padding:4px;vertical-align:middle;} .alphax-mis-excel .num{text-align:right;}</style>",
           "<table>"]
    for r in range(1, max_row + 1):
        out.append("<tr>")
        for c in range(1, max_col + 1):
            if (r, c) in skip:
                continue
            rs, cs = spans.get((r, c), (1, 1))
            attrs = ""
            if rs > 1: attrs += f" rowspan='{rs}'"
            if cs > 1: attrs += f" colspan='{cs}'"
            try:
                v = cells[r - 1][c - 1].get("v")
            except Exception:
                v = ""
            cls = ""
            if r in row_map and col_keys and c >= start_value_col:
                key = col_keys[c - start_value_col]
                v = fmt(values.get((r, key), 0.0))
                cls = " class='num'"
            out.append(f"<td{attrs}{cls}>{'' if v is None else v}</td>")
        out.append("</tr>")
    out.append("</table></div>")
    return "".join(out)
